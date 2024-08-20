''' 
    Gets filtered, alive/active devices from netconfig data dump and merge with relevant 
    data from google sheets. Insert final dataframe into IoT Asset Inventory Template 
    spreadsheet and generate the report.
'''

import os
import re
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from IPython.display import display

if __name__ == "__main__":

    # ingest the csv containing all the filtered, pingable devices from netconfig
    df = pd.read_csv('filtered_all.csv')

    f1 = open("inactive_hostnames.csv", 'r+')
    lines = f1.readline().split(', ')
    inactive = [host.replace('\'', '') for host in lines]

    # drop inactive devices from df based on hostname
    for name in inactive:
        df = df.drop(df[df['hostname'] == name].index)

    # rename 'hostname' column to 'device_name' for merge with google sheets dataframe
    df.rename(columns={'hostname': 'device_name'}, inplace=True)

    # ingest google sheets data
    df_google_sheet = pd.read_csv('google_sheets.csv')

    # drop columns with irrelevant data by index
    df_google_sheet.drop(
        df_google_sheet.columns[[0, 1, 2, 3, 41]], axis=1, inplace=True)

    # drop top two rows with irrelevant data
    df_google_sheet = df_google_sheet.iloc[2:]

    # set column labels to top row values
    df_google_sheet.columns = df_google_sheet.iloc[0]

    # remove the extra row with labels
    df_google_sheet = df_google_sheet.drop(
        index=df_google_sheet.iloc[0].name)

    # drop other unused columns
    drop_labels = ['Responsible for filling in yellow columns', 'Complete', 'Responsive to ping?', 'Description exclude', 'Criticality suggestion',
                   '4.2 Physical Location (Building or room)', 'Old Function data (moved here during cleanup)', 'MAC Vendor', '7.1 IP Address', 'Comments', 'Location (Based on netconfig data)', 'IP Address', 'Subnet Exclude', 'Device Name Exclude', 'Exclude?']

    for label in drop_labels:
        df_google_sheet = df_google_sheet.drop(label, axis=1)

    # rename df_google_sheet columns so they're easier to type
    df_google_sheet = df_google_sheet.rename(columns={'1.1 System Owner': 'system_owner', '2.1 System Name': 'system_name', '2.2 Device Name': 'device_name', '5.1 Manufacturer': 'manufacturer', '3.1 System Make': 'system_make', '3.2 System Model': 'system_model', '3.3 Relevent Specifications / Configurations': 'relevant_specs', '3.4 Unique Identifier': 'unique_identifier', '4.1 Function': 'function', '4.3 Criticality': 'criticality', '4.5 HVA System Association': 'hva_system_association',
                                             '5.2 Manufacturer Contact Information': 'manufacturer_contact', '5.3 Vendor': 'vendor', '5.4 Vendor Contact Information': 'vendor_contact', '5.5 Support Channels': 'support_channels', '6.1 Software Version Applied': 'software_version_applied', '6.2 Firmware Version Applied': 'firmware_version_applied', '6.3 Patch / Update Version Applied': 'patch_applied', '7.2 Port': 'port', '7.3 Integrations': 'integrations', '7.4 API': 'api', '7.5 Interconnective Communication Protocol': 'inter_comm_protocol', '8.1 Applied Security Controls': 'applied_security_controls', '8.1.1 Applied Security Control Comments': 'applied_security_comments'})

    # do left outer join on df and df_google_sheets
    df_final = df.merge(df_google_sheet, on='device_name', how='left')

    # put all data in 'support channels' column in quotes
    df_final['support_channels'] = [
        '\'' + str(value) + '\'' for value in df_final['support_channels']]

    # rearrange columns in df_final to match columns in iot_template.xlsx
    df_final = df_final.loc[:, ['system_owner', 'system_name', 'device_name', 'relevant_specs', 'unique_identifier', 'function', 'criticality', 'location', 'hva_system_association', 'manufacturer', 'manufacturer_contact', 'vendor',
                                'vendor_contact', 'support_channels', 'software_version_applied', 'firmware_version_applied', 'patch_applied', 'ip', 'port', 'integrations', 'api', 'inter_comm_protocol', 'applied_security_controls', 'applied_security_comments']]

    # add missing columns (from DoE Asset Inventory template) to df_final
    df_final.insert(3, 'device_make', '')
    df_final.insert(4, 'device_model', '')
    df_final.insert(9, 'critical_functions_documented', '')
    df_final.insert(10, 'related_systems_documented', '')
    df_final.insert(11, 'related_processes_documented', '')
    df_final.insert(12, 'related_assets_documented', '')
    df_final.insert(13, 'evaluation _of_attack_pathways', '')
    df_final.insert(14, 'prioritizing_risk_mitigation', '')
    df_final.insert(16, 'fisma_system_association', '')
    df_final.insert(17, 'fisma_system_name', '')
    df_final.insert(19, 'iot_doe_system_id', '')

    wb = openpyxl.load_workbook('iot_template.xlsx')
    ws = wb['Questions']
    ws.title = 'Questions'

    rows = dataframe_to_rows(df_final, index=False, header=False)

    for row_idx, row in enumerate(rows, 7):
        for col_idx, value in enumerate(row, 4):
            ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save('../iot_asset_inventory.xlsx')

    # for testing
    print(df_final.head(5))
